from PyQt5.QtWidgets import QCompleter, QMainWindow, QMessageBox, QApplication, QLabel, QLineEdit, QPushButton, QTableWidget, QFileDialog
from PyQt5 import uic, QtWidgets, QtGui, QtCore
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtGui import QIcon, QPixmap, QBrush, QColor, QFont
import sys, random, qdarkstyle
from openpyxl import load_workbook

class UI(QMainWindow):
    def __init__(self):
        super(UI, self).__init__()

        #Load UI
        uic.loadUi("NFL.ui",self) 

        #define widgets
        self.instructions = self.findChild(QPushButton, "pushButton")
        self.buttonDivisions = self.findChild(QPushButton, "pushButton_2")
        self.table = self.findChild(QTableWidget, "tableWidget")
        self.lineEdit = self.findChild(QLineEdit, "lineEdit")
        self.Guesses_Left_Lable = self.findChild(QLabel, "label_2")
        self.instructions.clicked.connect(self.popup)
        self.buttonDivisions.clicked.connect(self.popupDivisions)
        self.icon = QIcon()
        self.icon1 = QIcon()
        self.icon.addPixmap(QPixmap('info-icon.png'))
        self.instructions.setIcon(self.icon)
        iconDiv = QIcon()
        iconDiv.addPixmap(QPixmap('NFLICON.png'))  
        self.buttonDivisions.setIcon(iconDiv)      
        self.lineEdit.returnPressed.connect(self.guess)
        self.table.setRowHeight(0,0)
        self.table.verticalHeader().hide()
        self.table.setColumnWidth(0, 125)
        self.table.setColumnWidth(1, 125)
        self.table.setColumnWidth(2, 90)
        self.table.setColumnWidth(3, 90)
        self.table.setColumnWidth(4, 90)
        self.table.setColumnWidth(5, 90)
        self.table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff) 
        self.table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.setWindowTitle("NFL Wordle")
        self.icon.addPixmap(QPixmap('NFLICON.png'))
        self.setWindowIcon(self.icon)              
        self.show()
        #Get randomized answer player
        wb = load_workbook("NFLINFO.xlsx")
        self.ws = wb.active  
        column_names = self.ws['A'] 
        self.mylist = list() 
        i=0           
        for cell in column_names: 
            if(i>0):                      
                self.mylist.append(cell.value)
            i=i+1    
        #print(mylist)
        completer = QCompleter(self.mylist, self)
        completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)   
        completer.setFilterMode(Qt.MatchContains)
        self.lineEdit.setCompleter(completer)
        print(len(self.mylist))
        
              
        self.answerRow = random.randint(2,len(self.mylist)+1)
        spot = "A" + str(self.answerRow)
        self.AnswerName = self.ws[spot].value
        spot = "B" + str(self.answerRow)
        self.AnswerTeam = self.ws[spot].value
        spot = "C" + str(self.answerRow)
        self.AnswerDivision = self.ws[spot].value
        spot = "D" + str(self.answerRow)
        self.AnswerPosition = self.ws[spot].value
        spot = "E" + str(self.answerRow)
        self.AnswerAge = self.ws[spot].value
        spot = "F" + str(self.answerRow)
        self.AnswerNumber = self.ws[spot].value
        print('Answer is: ', self.AnswerName)
        print('At row', self.answerRow)                          

    #What happens when name is entered in lineEdit    
    def guess(self): 
        #clear lineEdit box for next guess
        QTimer.singleShot(0, self.lineEdit.clear)      
        #load in excel file
        wb = load_workbook("NFLINFO.xlsx")
        ws = wb.active       
        #search excel file in name column for a macthing name to the user input     
        for cell in ws['A']:           
            if(cell.value is not None):               
                # if there's a match then print it's row number (ignore case)
                if self.lineEdit.text().lower() == cell.value.lower():
                    self.excel_row = cell.row                   
                    #print('Row in excel file of guess: ',self.excel_row) 
        #popup for when user input dne in excel file                   
        try:
            self.excel_row           
        except AttributeError:           
            msgError = QMessageBox()
            msgError.setWindowTitle('Error') 
            self.icon.addPixmap(QPixmap("NFLICON.png"))
            msgError.setWindowIcon(self.icon)
            msgError.setText("Invalid Name")
            msgError.setIcon(QMessageBox.Warning)              
            self.lineEdit.clear()             
            x = msgError.exec_()
            return None 
        #Popup error when user hits enter but inputs nothing
        if(self.lineEdit.text()== '' or self.excel_row == ''):                     
            msgError = QMessageBox()
            msgError.setWindowTitle('Error') 
            self.icon.addPixmap(QPixmap("NFLICON.png"))
            msgError.setWindowIcon(self.icon)
            msgError.setText("Invalid Name:")
            msgError.setIcon(QMessageBox.Warning)              
            self.lineEdit.clear()             
            x = msgError.exec_()
            return None       

        self.GuessRow = random.randint(2,len(self.mylist))
        spot = "A" + str(self.excel_row)
        self.GuessName = self.ws[spot].value
        spot = "B" + str(self.excel_row)
        self.GuessTeam = self.ws[spot].value
        spot = "C" + str(self.excel_row)
        self.GuessDivision = self.ws[spot].value
        spot = "D" + str(self.excel_row)
        self.GuessPosition = self.ws[spot].value
        spot = "E" + str(self.excel_row)
        self.GuessAge = self.ws[spot].value
        spot = "F" + str(self.excel_row)
        self.GuessNumber = self.ws[spot].value
        print('guess is: ', self.GuessName)
        
        #Change label of remaining guesses
        Guesses_Remain = 8-self.table.rowCount()        
        if Guesses_Remain > 0:
            self.Guesses_Left_Lable.setText(str(Guesses_Remain))     
        #Add guess to table
        row_guess = ws[self.excel_row]
        row=self.table.rowCount()
        col=0       
        self.table.insertRow(self.table.rowCount())            
        for cell in row_guess:
            
            player_info = cell.value                      
            self.table.setItem(row, col, QtWidgets.QTableWidgetItem(str(player_info)))
            col = col+1 
                    
        current_row= self.table.rowCount()
        #Green to correct cells
        if(self.AnswerName == self.GuessName):
            self.table.item(current_row-1,0).setBackground(QtGui.QColor(116, 237, 130))
            self.table.item(current_row-1,0).setForeground(QBrush(QColor(25, 25, 25)))
        if(self.AnswerTeam == self.GuessTeam):
                self.table.item(current_row-1,1).setBackground(QtGui.QColor(116, 237, 130))
                self.table.item(current_row-1,1).setForeground(QBrush(QColor(25, 25, 25)))
        if(self.AnswerDivision == self.GuessDivision):
                self.table.item(current_row-1,2).setBackground(QtGui.QColor(116, 237, 130))
                self.table.item(current_row-1,2).setForeground(QBrush(QColor(25, 25, 25)))
        if(self.AnswerPosition == self.GuessPosition):
                self.table.item(current_row-1,3).setBackground(QtGui.QColor(116, 237, 130))
                self.table.item(current_row-1,3).setForeground(QBrush(QColor(25, 25, 25)))
        if(self.AnswerAge == self.GuessAge):
                self.table.item(current_row-1,4).setBackground(QtGui.QColor(116, 237, 130))
                self.table.item(current_row-1,4).setForeground(QBrush(QColor(25, 25, 25)))            
        if(self.AnswerNumber == self.GuessNumber):
                self.table.item(current_row-1,5).setBackground(QtGui.QColor(116, 237, 130))
                self.table.item(current_row-1,5).setForeground(QBrush(QColor(25, 25, 25)))
        #Yellow for AFC Conference
        if(self.AnswerDivision != self.GuessDivision):
            if(self.AnswerDivision == 'AFC North' or self.AnswerDivision == 'AFC South' or self.AnswerDivision == 'AFC East' or self.AnswerDivision == 'AFC West'):
                if(self.GuessDivision == 'AFC North' or self.GuessDivision == 'AFC South' or self.GuessDivision == 'AFC East' or self.GuessDivision == 'AFC West'):
                    self.table.item(current_row-1,2).setBackground(QtGui.QColor(252, 255, 99))
                    self.table.item(current_row-1,2).setForeground(QBrush(QColor(25, 25, 25)))
        #Yellow for NFC Conference
        if(self.AnswerDivision != self.GuessDivision):
            if(self.AnswerDivision == 'NFC North' or self.AnswerDivision == 'NFC South' or self.AnswerDivision == 'NFC East' or self.AnswerDivision == 'NFC West'):
                if(self.GuessDivision == 'NFC North' or self.GuessDivision == 'NFC South' or self.GuessDivision == 'NFC East' or self.GuessDivision == 'NFC West'):
                    self.table.item(current_row-1,2).setBackground(QtGui.QColor(252, 255, 99))
                    self.table.item(current_row-1,2).setForeground(QBrush(QColor(25, 25, 25)))            
        #Yellow for offense match
        if(self.AnswerPosition != self.GuessPosition):
            if(self.AnswerPosition == 'QB' or self.AnswerPosition == 'RB' or self.AnswerPosition == 'WR' or self.AnswerPosition == 'TE' or self.AnswerPosition == 'OT' or self.AnswerPosition == 'G' or self.AnswerPosition == 'C' or self.AnswerPosition == 'FB'):
                if(self.GuessPosition == 'QB' or self.GuessPosition == 'RB' or self.GuessPosition == 'WR' or self.GuessPosition == 'TE' or self.GuessPosition == 'OT' or self.GuessPosition == 'G' or self.GuessPosition == 'C' or self.GuessPosition == 'FB'):
                        print(self.AnswerPosition)
                        print(self.GuessPosition)
                        self.table.item(current_row-1,3).setBackground(QtGui.QColor(252, 255, 99))
                        self.table.item(current_row-1,3).setForeground(QBrush(QColor(25, 25, 25)))
        #Yellow for defense match
        if(self.AnswerPosition != self.GuessPosition):
            if(self.AnswerPosition == 'LB' or self.AnswerPosition == 'S' or self.AnswerPosition == 'DE' or self.AnswerPosition == 'DT' or self.AnswerPosition == 'CB'):
                if(self.GuessPosition == 'LB' or self.GuessPosition == 'S' or self.GuessPosition == 'DE' or self.GuessPosition == 'DT' or self.GuessPosition == 'CB'):
                        print(self.AnswerPosition)
                        print(self.GuessPosition)                       
                        self.table.item(current_row-1,3).setBackground(QtGui.QColor(252, 255, 99))
                        self.table.item(current_row-1,3).setForeground(QBrush(QColor(25, 25, 25)))
        #Yellow for age guess over
        if(self.AnswerAge < self.GuessAge):
            self.table.item(current_row-1,4).setBackground(QtGui.QColor(252, 255, 99))
            self.table.item(current_row-1,4).setForeground(QBrush(QColor(25, 25, 25)))
        #yellow for jersey number
        if(self.AnswerNumber < self.GuessNumber):
            self.table.item(current_row-1,5).setBackground(QtGui.QColor(252, 255, 99))
            self.table.item(current_row-1,5).setForeground(QBrush(QColor(25, 25, 25)))
        
        if Guesses_Remain == 0 and self.AnswerName != self.GuessName:
            msgLose = QMessageBox()
            msgLose.setWindowTitle('Game Over') 
            self.icon.addPixmap(QPixmap('NFLICON.png'))
            msgLose.setWindowIcon(self.icon)
            msgLose.setText("You ran out of guesses! The correct answer was %s " %(self.AnswerName))
            #msg.setIcon(QMessageBox.Warning)                                     
            x = msgLose.exec_()
            while(self.table.rowCount() > 1):
                self.table.removeRow(1)
            self.Guesses_Left_Lable.setText(str(8))
            self.answerRow = random.randint(2,len(self.mylist))
            spot = "A" + str(self.answerRow)
            self.AnswerName = self.ws[spot].value
            spot = "B" + str(self.answerRow)
            self.AnswerTeam = self.ws[spot].value
            spot = "C" + str(self.answerRow)
            self.AnswerDivision = self.ws[spot].value
            spot = "D" + str(self.answerRow)
            self.AnswerPosition = self.ws[spot].value
            spot = "E" + str(self.answerRow)
            self.AnswerAge = self.ws[spot].value
            spot = "F" + str(self.answerRow)
            self.AnswerNumber = self.ws[spot].value
            print(self.AnswerName)           
            print(self.answerRow)
            QTimer.singleShot(0, self.lineEdit.clear)
            return None 
        else:
            self.Guesses_Left_Lable.setText(str(Guesses_Remain))
        #Popup for when guess == answer              
        if(self.GuessName == self.AnswerName):
            msgAnswer = QMessageBox()
            msgAnswer.setWindowTitle('Success!')            
            self.icon.addPixmap(QPixmap('NFLICON.png'))
            msgAnswer.setWindowIcon(self.icon)           
            msgAnswer.setText("You guessed correctly and only took %s guesses! " %(8-Guesses_Remain))                                                                
            x = msgAnswer.exec_()
            while(self.table.rowCount() > 1):
                self.table.removeRow(1)
            self.Guesses_Left_Lable.setText(str(8))
            self.answerRow = random.randint(2,len(self.mylist))
            spot = "A" + str(self.answerRow)
            self.AnswerName = self.ws[spot].value
            spot = "B" + str(self.answerRow)
            self.AnswerTeam = self.ws[spot].value
            spot = "C" + str(self.answerRow)
            self.AnswerDivision = self.ws[spot].value
            spot = "D" + str(self.answerRow)
            self.AnswerPosition = self.ws[spot].value
            spot = "E" + str(self.answerRow)
            self.AnswerAge = self.ws[spot].value
            spot = "F" + str(self.answerRow)
            self.AnswerNumber = self.ws[spot].value
            print(self.AnswerName)           
            print(self.answerRow) 
            QTimer.singleShot(0, self.lineEdit.clear)                     
            return None 
        self.excel_row = ''                                   

    #command for instructions button   
    def popup(self):
        msgIns = QMessageBox()             
        msgIns.setWindowTitle('Instructions') 
        self.icon1.addPixmap(QPixmap('info-icon.png'))
        msgIns.setWindowIcon(self.icon1)
        msgIns.setInformativeText("Guess the NFL player in as few of tries as possible. After guesses the colors of each cell will indicate how close your guess was. Green means correct. Yellow varies: For position yellow indicates correct side of the ball (offense/defense), while yellow for division means correct conference (AFC/NFC). For age and number yellow means your guess was too high. Grey means incorrect or guess was too low.")                                  
        x = msgIns.exec_()              
    def popupDivisions(self):
        msgDiv = QMessageBox()
        msgDiv.setWindowTitle('NFL Division/Conference Layout')
        icon3 = QIcon()
        icon3.addPixmap(QPixmap('NFLICON.png'))
        msgDiv.setWindowIcon(icon3)
        msgDiv.setInformativeText("AFC North: Ravens, Bengals, Browns, Steelers. <br> AFC East: Bills, Dolphins, Patriots, Jets. <br> AFC South: Texans, Colts, Titans, Jaguars. <br> AFC West: Broncos, Chiefs, Raiders, Chargers. <br><br> NFC North: Bears, Lions, Packers, Vikings. <br> NFC East: Cowboys, Giants, Eagles, Commanders. <br> NFC South: Falcons, Panthers, Saints, Buccaneers. <br> NFC West: Cardinals, Rams, 49ers, Seahawks")
        msgDiv.setFont(QFont('Times',12))
        x= msgDiv.exec_()
        
#initialize app
app = QApplication(sys.argv)
app.setStyleSheet(qdarkstyle.load_stylesheet())
UIWindow = UI()
app.exec_()